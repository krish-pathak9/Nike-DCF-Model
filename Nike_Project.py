import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, numbers, Border, Side
from openpyxl import load_workbook

# === Load Excel Financials ===
cf = pd.read_excel("Nike_Financials.xlsx", sheet_name="Nike_CF", skiprows=10, index_col=0)
bs = pd.read_excel("Nike_Financials.xlsx", sheet_name="Nike_BS", skiprows=10, index_col=0)
is_ = pd.read_excel("Nike_Financials.xlsx", sheet_name="Nike_IS", skiprows=10, index_col=0)

# === Extract key inputs ===
revenue_hist = is_.loc[is_.index.str.contains("Revenues", case=False)].iloc[0]
cogs_hist = is_.loc[is_.index.str.contains("Cost of sales", case=False)].iloc[0]
sgna_hist = is_.loc[is_.index.str.contains("Total selling & administrative expense", case=False)].iloc[0]
dna_hist = cf[cf.index.str.contains("D&A", case=False)].squeeze()
dna_hist = dna_hist[revenue_hist.index]
capex_hist = cf.loc["Additions to property, plant & equipment"]
tax_rate = 0.25

# === Driver ratios as % of revenue ===
cogs_pct = float((cogs_hist / revenue_hist).mean())
sgna_pct = float((sgna_hist / revenue_hist).mean())
dna_pct = float((dna_hist / revenue_hist).mean())
capex_pct = float((capex_hist / revenue_hist).mean())

# === Projection Assumptions ===
years_proj = [2025, 2026, 2027, 2028, 2029]
growth_rate = 0.02
revenue_base = revenue_hist.iloc[0]
revenue_proj = [revenue_base * ((1 + growth_rate) ** i) for i in range(1, 6)]
cogs_proj = [r * cogs_pct for r in revenue_proj]
sgna_proj = [r * sgna_pct for r in revenue_proj]
dna_proj = [r * dna_pct for r in revenue_proj]
capex_proj = [r * capex_pct for r in revenue_proj]
ebit_proj = [r - c - s for r, c, s in zip(revenue_proj, cogs_proj, sgna_proj)]
tax_proj = [e * tax_rate for e in ebit_proj]
tax_ebit_proj = [e - t for e, t in zip(ebit_proj, tax_proj)]
ufcf_proj = [te + d - x for te, d, x in zip(tax_ebit_proj, dna_proj, capex_proj)]

forecast_df = pd.DataFrame({
    "Year": years_proj,
    "Revenue": revenue_proj,
    "COGS": cogs_proj,
    "SG&A": sgna_proj,
    "EBIT": ebit_proj,
    "Taxes (25%)": tax_proj,
    "Tax-Affected EBIT": tax_ebit_proj,
    "D&A": dna_proj,
    "CapEx": capex_proj,
    "Unlevered FCF": ufcf_proj
})

# === DCF Valuation ===
wacc = 0.09
tv_growth = 0.03
discount_factors = [(1 + wacc) ** (i + 1) for i in range(len(ufcf_proj))]
discounted_fcfs = [fcf / df for fcf, df in zip(ufcf_proj, discount_factors)]
terminal_fcf = ufcf_proj[-1] * (1 + tv_growth)
terminal_value = terminal_fcf / (wacc - tv_growth)
discounted_terminal_value = terminal_value / discount_factors[-1]
enterprise_value = sum(discounted_fcfs) + discounted_terminal_value

# === Summary Tables ===
dcf_summary_df = pd.DataFrame([
    ["Sum of Discounted FCFs", sum(discounted_fcfs)],
    ["Discounted Terminal Value", discounted_terminal_value],
    ["Enterprise Value", enterprise_value],
    ["Excel File", "Nike_Forecast_DCF_Model.xlsx"]
], columns=["Component", "Value"])

assumptions_df = pd.DataFrame({
    "Assumption": [
        "Revenue Growth",
        "COGS as % of Sales",
        "SG&A as % of Sales",
        "D&A as % of Sales",
        "CapEx as % of Sales",
        "Tax Rate",
        "WACC",
        "Terminal Growth"
    ],
    "Value": [
        f"{growth_rate * 100:.1f}%",
        f"{cogs_pct * 100:.1f}%",
        f"{sgna_pct * 100:.1f}%",
        f"{dna_pct * 100:.1f}%",
        f"{capex_pct * 100:.1f}%",
        f"{tax_rate * 100:.1f}%",
        f"{wacc * 100:.1f}%",
        f"{tv_growth * 100:.1f}%"
    ],
    "Explanation": [
        "Annual growth rate applied to revenue base",
        "Historical average cost of goods sold as % of sales",
        "Selling, general & admin costs as % of sales",
        "Depreciation and amortization relative to sales",
        "Capital expenditure forecasted as % of sales",
        "Applied to EBIT to calculate tax expense",
        "Discount rate used to value cash flows",
        "Used to estimate terminal value beyond forecast"
    ]
})

# === Export to Excel ===
output_file = "Nike_Forecast_DCF_Model.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    forecast_df.to_excel(writer, sheet_name="Projected FCF", index=False)
    assumptions_df.to_excel(writer, sheet_name="Forecast Assumptions", index=False)
    dcf_summary_df.to_excel(writer, sheet_name="DCF Summary", index=False)

# === Excel Formatting ===
wb = load_workbook(output_file)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                label = str(ws.cell(row=cell.row, column=1).value)
                if label and any(word in label for word in ["Rate", "%", "Growth"]):
                    cell.number_format = "0.00%"
                elif "Year" in label:
                    cell.number_format = "0"
                else:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = Alignment(horizontal="right")
                cell.border = thin_border

    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

wb.save(output_file)
print(f"✅ Simplified and professional DCF model saved to Excel: {output_file}")